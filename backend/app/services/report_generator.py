from io import BytesIO
from typing import List
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, HRFlowable
from app.models import VisitReport, ExpenseReport

PRIMARY = "1E40AF"
LIGHT = "EFF6FF"
HEADER_COLOR = colors.HexColor("#1E40AF")
ROW_ALT = colors.HexColor("#EFF6FF")

_thin = Side(border_style="thin", color="CBD5E1")
_border = Border(left=_thin, right=_thin, top=_thin, bottom=_thin)


def _apply_header_row(ws, row: int, headers: list, col_widths: list):
    fill = PatternFill(fill_type="solid", fgColor=PRIMARY)
    font = Font(bold=True, color="FFFFFF", size=10)
    for col, (header, width) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _border
        ws.column_dimensions[get_column_letter(col)].width = width


def _apply_data_row(ws, row: int, values: list, alt: bool = False):
    fill = PatternFill(fill_type="solid", fgColor=LIGHT) if alt else None
    for col, value in enumerate(values, 1):
        cell = ws.cell(row=row, column=col, value=value)
        cell.alignment = Alignment(vertical="center", wrap_text=True)
        cell.border = _border
        if fill:
            cell.fill = fill


def generate_visit_excel(reports: List[VisitReport], rep_name: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Visit Report"
    ws.row_dimensions[1].height = 30
    ws.row_dimensions[3].height = 20

    ws.merge_cells("A1:F1")
    title_cell = ws["A1"]
    title_cell.value = "OREXIS PHARMA — Doctor Visit Report"
    title_cell.font = Font(bold=True, size=14, color="1E40AF")
    title_cell.alignment = Alignment(horizontal="center", vertical="center")

    ws["A2"] = f"Representative: {rep_name}"
    ws["A2"].font = Font(bold=True, size=10)

    headers = ["Date", "Area", "Doctor Name", "Speciality", "Visited", "Reason (if not visited)"]
    widths = [14, 16, 22, 18, 10, 32]
    _apply_header_row(ws, 4, headers, widths)

    row_num = 5
    for report in reports:
        for item in report.items:
            _apply_data_row(ws, row_num, [
                str(report.report_date),
                report.area.name,
                item.doctor.name,
                item.doctor.speciality or "",
                "Yes" if item.visited else "No",
                item.note or "",
            ], alt=(row_num % 2 == 0))
            row_num += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_expense_excel(reports: List[ExpenseReport], rep_name: str) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expense Report"

    ws.merge_cells("A1:E1")
    ws["A1"].value = "OREXIS PHARMA — Travel Expense Report"
    ws["A1"].font = Font(bold=True, size=14, color="1E40AF")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws["A2"] = f"Representative: {rep_name}"
    ws["A2"].font = Font(bold=True, size=10)

    headers = ["Date", "Area", "Description", "Amount (₹)", "Status"]
    widths = [14, 16, 30, 14, 14]
    _apply_header_row(ws, 4, headers, widths)

    row_num = 5
    for report in reports:
        for item in report.items:
            _apply_data_row(ws, row_num, [
                str(report.report_date),
                report.area.name,
                item.description,
                float(item.amount),
                report.status.value.capitalize(),
            ], alt=(row_num % 2 == 0))
            row_num += 1
        ws.cell(row=row_num, column=3, value="TOTAL").font = Font(bold=True)
        ws.cell(row=row_num, column=4, value=float(report.total_amount)).font = Font(bold=True)
        row_num += 1

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def generate_visit_pdf(reports: List[VisitReport], rep_name: str) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=16, textColor=HEADER_COLOR, spaceAfter=4)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontSize=10, spaceAfter=8)

    elements.append(Paragraph("OREXIS PHARMA — Doctor Visit Report", title_style))
    elements.append(Paragraph(f"Representative: {rep_name}", sub_style))
    elements.append(HRFlowable(width="100%", color=HEADER_COLOR, spaceAfter=8))

    data = [["Date", "Area", "Doctor Name", "Speciality", "Visited", "Reason (if not visited)"]]
    for report in reports:
        for item in report.items:
            data.append([
                str(report.report_date),
                report.area.name,
                item.doctor.name,
                item.doctor.speciality or "",
                "Yes" if item.visited else "No",
                item.note or "",
            ])

    col_widths = [1.1 * inch, 1.3 * inch, 1.8 * inch, 1.5 * inch, 0.7 * inch, 2.6 * inch]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HEADER_COLOR),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ROW_ALT]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("PADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()


def generate_expense_pdf(reports: List[ExpenseReport], rep_name: str) -> bytes:
    buf = BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), topMargin=0.5 * inch, bottomMargin=0.5 * inch)
    styles = getSampleStyleSheet()
    elements = []

    title_style = ParagraphStyle("title", parent=styles["Heading1"], fontSize=16, textColor=HEADER_COLOR, spaceAfter=4)
    sub_style = ParagraphStyle("sub", parent=styles["Normal"], fontSize=10, spaceAfter=8)

    elements.append(Paragraph("OREXIS PHARMA — Travel Expense Report", title_style))
    elements.append(Paragraph(f"Representative: {rep_name}", sub_style))
    elements.append(HRFlowable(width="100%", color=HEADER_COLOR, spaceAfter=8))

    data = [["Date", "Area", "Description", "Amount (₹)", "Status"]]
    for report in reports:
        for item in report.items:
            data.append([
                str(report.report_date),
                report.area.name,
                item.description,
                f"₹{float(item.amount):,.2f}",
                report.status.value.capitalize(),
            ])
        data.append(["", "", "TOTAL", f"₹{float(report.total_amount):,.2f}", ""])

    col_widths = [1.1 * inch, 1.3 * inch, 3.0 * inch, 1.2 * inch, 1.0 * inch]
    table = Table(data, colWidths=col_widths, repeatRows=1)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), HEADER_COLOR),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, ROW_ALT]),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#CBD5E1")),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("PADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(table)
    doc.build(elements)
    return buf.getvalue()
